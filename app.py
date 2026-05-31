import os
import sqlite3
import tempfile
from datetime import datetime
from functools import wraps
from io import BytesIO

from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash


APP_DIR = os.path.dirname(os.path.abspath(__file__))

if os.name == "nt":
    default_db_path = os.path.join(tempfile.gettempdir(), "axystry_gastos.db")
else:
    default_db_path = os.path.join(APP_DIR, ".runtime", "gastos.db")

DB_PATH = os.environ.get("GASTOS_DB_PATH", default_db_path)

db_dir = os.path.dirname(DB_PATH)
if db_dir:
    os.makedirs(db_dir, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-esta-clave-secreta")

# Usuario inicial del sistema
USUARIO_ADMIN = os.environ.get("GASTOS_USER", "admin")
PASSWORD_ADMIN_HASH = generate_password_hash(os.environ.get("GASTOS_PASSWORD", "12345"))

CATEGORIAS = [
    "Transporte",
    "Alimentación",
    "Software",
    "Publicidad",
    "Oficina",
    "Equipos",
    "Servicios",
    "Otros",
]

MONEDAS = ["CLP", "USD", "UYU", "ARS", "BRL", "EUR"]

ESTADO_ACTIVO = "Activo"
ESTADO_ANULADO = "Anulado"


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def columna_existe(conn, tabla, columna):
    columnas = [col[1] for col in conn.execute(f"PRAGMA table_info({tabla})").fetchall()]
    return columna in columnas


def init_db():
    conn = get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            responsable TEXT NOT NULL DEFAULT '',
            fecha TEXT NOT NULL,
            monto REAL NOT NULL,
            moneda TEXT NOT NULL,
            categoria TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'Activo',
            fecha_anulacion TEXT,
            creado_en TEXT NOT NULL,
            actualizado_en TEXT
        )
    """)

    # Migraciones automáticas para bases ya existentes.
    # Esto permite actualizar el sistema sin perder los gastos ya cargados.
    if not columna_existe(conn, "gastos", "responsable"):
        conn.execute("ALTER TABLE gastos ADD COLUMN responsable TEXT NOT NULL DEFAULT ''")

    if not columna_existe(conn, "gastos", "estado"):
        conn.execute("ALTER TABLE gastos ADD COLUMN estado TEXT NOT NULL DEFAULT 'Activo'")

    if not columna_existe(conn, "gastos", "fecha_anulacion"):
        conn.execute("ALTER TABLE gastos ADD COLUMN fecha_anulacion TEXT")

    if not columna_existe(conn, "gastos", "actualizado_en"):
        conn.execute("ALTER TABLE gastos ADD COLUMN actualizado_en TEXT")

    conn.commit()
    conn.close()


def login_requerido(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("logueado"):
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def obtener_gastos(incluir_anulados=True):
    conn = get_db_connection()

    if incluir_anulados:
        gastos = conn.execute("""
            SELECT id, responsable, fecha, monto, moneda, categoria, descripcion,
                   estado, fecha_anulacion, creado_en, actualizado_en
            FROM gastos
            ORDER BY
                CASE WHEN estado = 'Anulado' THEN 1 ELSE 0 END,
                fecha DESC,
                id DESC
        """).fetchall()
    else:
        gastos = conn.execute("""
            SELECT id, responsable, fecha, monto, moneda, categoria, descripcion,
                   estado, fecha_anulacion, creado_en, actualizado_en
            FROM gastos
            WHERE estado = 'Activo'
            ORDER BY fecha DESC, id DESC
        """).fetchall()

    conn.close()
    return gastos


def obtener_gasto_por_id(gasto_id):
    conn = get_db_connection()
    gasto = conn.execute("""
        SELECT id, responsable, fecha, monto, moneda, categoria, descripcion,
               estado, fecha_anulacion, creado_en, actualizado_en
        FROM gastos
        WHERE id = ?
    """, (gasto_id,)).fetchone()
    conn.close()
    return gasto


def obtener_resumen_por_categoria():
    conn = get_db_connection()
    resumen = conn.execute("""
        SELECT categoria, SUM(monto) AS total
        FROM gastos
        WHERE estado = 'Activo'
        GROUP BY categoria
        ORDER BY total DESC
    """).fetchall()
    conn.close()
    return resumen


def obtener_resumen_por_responsable():
    conn = get_db_connection()
    resumen = conn.execute("""
        SELECT
            CASE
                WHEN TRIM(responsable) = '' THEN 'Sin responsable'
                ELSE responsable
            END AS responsable,
            SUM(monto) AS total
        FROM gastos
        WHERE estado = 'Activo'
        GROUP BY CASE
            WHEN TRIM(responsable) = '' THEN 'Sin responsable'
            ELSE responsable
        END
        ORDER BY total DESC
    """).fetchall()
    conn.close()
    return resumen


def validar_datos_gasto(responsable, fecha, monto, moneda, categoria, descripcion):
    errores = []

    if not responsable:
        errores.append("El responsable es obligatorio.")

    if not fecha:
        errores.append("La fecha es obligatoria.")

    try:
        monto_num = float(monto.replace(",", "."))
        if monto_num <= 0:
            errores.append("El monto debe ser mayor a cero.")
    except ValueError:
        errores.append("El monto debe ser numérico.")
        monto_num = None

    if moneda not in MONEDAS:
        errores.append("La moneda seleccionada no es válida.")

    if categoria not in CATEGORIAS:
        errores.append("La categoría seleccionada no es válida.")

    if not descripcion:
        errores.append("La descripción es obligatoria.")

    return errores, monto_num


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        password = request.form.get("password", "").strip()

        if usuario == USUARIO_ADMIN and check_password_hash(PASSWORD_ADMIN_HASH, password):
            session["logueado"] = True
            session["usuario"] = usuario
            return redirect(url_for("index"))

        flash("Usuario o contraseña incorrectos.", "error")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/", methods=["GET", "POST"])
@login_requerido
def index():
    if request.method == "POST":
        responsable = request.form.get("responsable", "").strip()
        fecha = request.form.get("fecha", "").strip()
        monto = request.form.get("monto", "").strip()
        moneda = request.form.get("moneda", "").strip()
        categoria = request.form.get("categoria", "").strip()
        descripcion = request.form.get("descripcion", "").strip()

        errores, monto_num = validar_datos_gasto(
            responsable, fecha, monto, moneda, categoria, descripcion
        )

        if errores:
            for error in errores:
                flash(error, "error")
        else:
            conn = get_db_connection()
            conn.execute("""
                INSERT INTO gastos (
                    responsable, fecha, monto, moneda, categoria, descripcion,
                    estado, creado_en, actualizado_en
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                responsable,
                fecha,
                monto_num,
                moneda,
                categoria,
                descripcion,
                ESTADO_ACTIVO,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ))
            conn.commit()
            conn.close()
            flash("Gasto cargado correctamente.", "success")
            return redirect(url_for("index"))

    gastos = obtener_gastos(incluir_anulados=True)
    resumen = obtener_resumen_por_categoria()
    resumen_responsable = obtener_resumen_por_responsable()

    labels = [row["categoria"] for row in resumen]
    values = [round(row["total"], 2) for row in resumen]

    labels_responsable = [row["responsable"] for row in resumen_responsable]
    values_responsable = [round(row["total"], 2) for row in resumen_responsable]

    total_general = sum(values)

    return render_template(
        "index.html",
        gastos=gastos,
        categorias=CATEGORIAS,
        monedas=MONEDAS,
        labels=labels,
        values=values,
        labels_responsable=labels_responsable,
        values_responsable=values_responsable,
        total_general=total_general,
    )


@app.route("/editar/<int:gasto_id>", methods=["GET", "POST"])
@login_requerido
def editar_gasto(gasto_id):
    gasto = obtener_gasto_por_id(gasto_id)

    if not gasto:
        flash("El gasto solicitado no existe.", "error")
        return redirect(url_for("index"))

    if gasto["estado"] == ESTADO_ANULADO:
        flash("No se puede editar un gasto anulado.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        responsable = request.form.get("responsable", "").strip()
        fecha = request.form.get("fecha", "").strip()
        monto = request.form.get("monto", "").strip()
        moneda = request.form.get("moneda", "").strip()
        categoria = request.form.get("categoria", "").strip()
        descripcion = request.form.get("descripcion", "").strip()

        errores, monto_num = validar_datos_gasto(
            responsable, fecha, monto, moneda, categoria, descripcion
        )

        if errores:
            for error in errores:
                flash(error, "error")
        else:
            conn = get_db_connection()
            conn.execute("""
                UPDATE gastos
                SET responsable = ?,
                    fecha = ?,
                    monto = ?,
                    moneda = ?,
                    categoria = ?,
                    descripcion = ?,
                    actualizado_en = ?
                WHERE id = ?
            """, (
                responsable,
                fecha,
                monto_num,
                moneda,
                categoria,
                descripcion,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                gasto_id,
            ))
            conn.commit()
            conn.close()

            flash("Gasto actualizado correctamente.", "success")
            return redirect(url_for("index"))

    return render_template(
        "editar.html",
        gasto=gasto,
        categorias=CATEGORIAS,
        monedas=MONEDAS,
    )


@app.route("/anular/<int:gasto_id>", methods=["POST"])
@login_requerido
def anular_gasto(gasto_id):
    gasto = obtener_gasto_por_id(gasto_id)

    if not gasto:
        flash("El gasto solicitado no existe.", "error")
        return redirect(url_for("index"))

    if gasto["estado"] == ESTADO_ANULADO:
        flash("El gasto ya estaba anulado.", "error")
        return redirect(url_for("index"))

    conn = get_db_connection()
    conn.execute("""
        UPDATE gastos
        SET estado = ?,
            fecha_anulacion = ?,
            actualizado_en = ?
        WHERE id = ?
    """, (
        ESTADO_ANULADO,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        gasto_id,
    ))
    conn.commit()
    conn.close()

    flash("Gasto anulado correctamente.", "success")
    return redirect(url_for("index"))


@app.route("/exportar")
@login_requerido
def exportar_excel():
    # El Excel exporta solo gastos activos.
    # Los anulados se mantienen en base para historial, pero no afectan reportes ni totales.
    gastos = obtener_gastos(incluir_anulados=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    encabezados = [
        "ID",
        "Responsable",
        "Fecha",
        "Monto",
        "Moneda",
        "Categoría",
        "Descripción",
        "Estado",
        "Creado en",
        "Actualizado en",
    ]
    ws.append(encabezados)

    for gasto in gastos:
        ws.append([
            gasto["id"],
            gasto["responsable"],
            gasto["fecha"],
            gasto["monto"],
            gasto["moneda"],
            gasto["categoria"],
            gasto["descripcion"],
            gasto["estado"],
            gasto["creado_en"],
            gasto["actualizado_en"],
        ])

    # ==========================
    # ESTILO PROFESIONAL EXCEL
    # ==========================
    azul_header = "1E3A8A"
    blanco = "FFFFFF"
    gris_borde = "CBD5E1"
    gris_suave = "F8FAFC"

    header_fill = PatternFill("solid", fgColor=azul_header)
    header_font = Font(color=blanco, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color=gris_borde),
        right=Side(style="thin", color=gris_borde),
        top=Side(style="thin", color=gris_borde),
        bottom=Side(style="thin", color=gris_borde),
    )

    # Encabezados: barra azul con letras blancas
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Filas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # Sombreado alterno para mejorar lectura
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=gris_suave)

    # Formatos específicos
    for row in range(2, ws.max_row + 1):
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row}"].number_format = "yyyy-mm-dd"
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"G{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Congelar encabezado y activar filtros
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # Altura de encabezado
    ws.row_dimensions[1].height = 24

    # Anchos controlados por columna
    anchos = {
        "A": 8,    # ID
        "B": 20,   # Responsable
        "C": 14,   # Fecha
        "D": 14,   # Monto
        "E": 12,   # Moneda
        "F": 18,   # Categoría
        "G": 36,   # Descripción
        "H": 14,   # Estado
        "I": 20,   # Creado en
        "J": 20,   # Actualizado en
    }

    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    # Autofit simple con límites razonables
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        calculated_width = min(max(max_length + 3, anchos.get(column_letter, 12)), 42)
        ws.column_dimensions[column_letter].width = calculated_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"gastos_empresa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


init_db()


if __name__ == "__main__":
    app.run(debug=True)
